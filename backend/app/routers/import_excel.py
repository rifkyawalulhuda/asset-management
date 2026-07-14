import io
import os
from fastapi import APIRouter, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
from fastapi import Depends
from app.database import get_db
from app.utils.excel_importer import import_fa_sheet, import_approval_sheet
import openpyxl

router = APIRouter(prefix="/import", tags=["import"])


@router.post("/excel")
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")

    results = {}

    if "FA_2026" in wb.sheetnames:
        n = import_fa_sheet(wb["FA_2026"], 2026, db)
        results["FA_2026"] = n

    if "FA_2025" in wb.sheetnames:
        n = import_fa_sheet(wb["FA_2025"], 2025, db)
        results["FA_2025"] = n

    db.commit()

    for sheet_name in ["List Approval 2022", "List Approval 2021"]:
        if sheet_name in wb.sheetnames:
            year = int(sheet_name.split()[-1])
            n = import_approval_sheet(wb[sheet_name], year, db)
            results[sheet_name] = n
            db.commit()

    total = sum(results.values())
    message = f"Imported {total} records: " + ", ".join(f"{k}: {v}" for k, v in results.items())

    return {"message": message, "details": results}
