import io
from typing import Optional
from datetime import datetime, date
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models.fixed_asset import FixedAsset
from app.models.depreciation_monthly import DepreciationMonthly
from app.models.acquisition_disposal import AcquisitionDisposal
from app.models.planned_asset import PlannedAsset

router = APIRouter(prefix="/export", tags=["export"])

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_KEYS = ["jan", "feb", "mar", "apr", "may", "jun",
              "jul", "aug", "sep", "oct", "nov", "dec"]

# ── Style helpers ──────────────────────────────────────────────────────────────

def make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def make_font(bold=False, color="000000", size=9) -> Font:
    return Font(bold=bold, color=color, size=size)

def make_border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def apply_header(ws, row: int, cols: list, fill_hex: str, font_color="FFFFFF"):
    fill = make_fill(fill_hex)
    font = make_font(bold=True, color=font_color)
    border = make_border()
    for ci, value in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=ci, value=value)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def apply_row(ws, row: int, values: list, fill_hex: str = None, number_cols: list = None):
    fill = make_fill(fill_hex) if fill_hex else None
    border = make_border()
    for ci, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=ci, value=value)
        cell.border = border
        if fill:
            cell.fill = fill
        if number_cols and ci in number_cols and value is not None:
            try:
                cell.value = float(value)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            except (TypeError, ValueError):
                pass

def auto_col_width(ws, min_width=8, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


# ── Sheet builders ─────────────────────────────────────────────────────────────

def build_asset_list(wb: Workbook, assets: list, year_ref: int):
    ws = wb.create_sheet("Asset List")
    ws.sheet_properties.tabColor = "1E3A5F"
    ws.freeze_panes = "A2"

    headers = [
        "No.", "Site", "Job", "Account No", "Category", "Asset No",
        "Fixed Asset No (AX)", "Purchase Date", "Group", "Voucher No",
        "Name of Fixed Asset", "Maker/Type/Location", "Capacity/Size/User",
        "Year", "Police No", "Machine No", "Chasis No", "Qty", "Valas",
        "Purchase Price", "Monthly Depreciation", "Period Total (mo)",
        "Acc Period Prev Year", "Period Yearly", "Period Until Year", "Period Remain",
        "Acc Depreciation Prev", "NBV Prev",
        "Dep Expense Current", "Acc Depreciation Curr", "NBV Current",
        "Additional", "Disposals", "Condition", "Photo Status", "Remark"
    ]
    apply_header(ws, 1, headers, "1E3A5F")
    # Number format columns (1-indexed): Purchase Price=20, Monthly Dep=21,
    # Acc Period cols=23-26 (numbers but small), Acc Dep Prev=27, NBV Prev=28,
    # Dep Exp=29, Acc Dep Curr=30, NBV Curr=31
    num_cols = {20, 21, 27, 28, 29, 30, 31}

    for ri, asset in enumerate(assets, start=2):
        row_fill = "F0F4FF" if ri % 2 == 0 else None
        values = [
            asset.no,
            asset.site_location,
            asset.job,
            asset.account_no,
            asset.category,
            asset.asset_no,
            asset.fixed_asset_number_ax,
            asset.purchase_date,
            asset.group_name,
            asset.voucher_no,
            asset.name,
            asset.maker_type_location,
            asset.capacity_size_user,
            asset.year,
            asset.police_no,
            asset.machine_no,
            asset.chasis_no,
            asset.quantity,
            asset.valas,
            float(asset.purchase_price) if asset.purchase_price else None,
            float(asset.monthly_depreciation) if asset.monthly_depreciation else None,
            asset.depreciation_period_total,
            float(asset.dep_period_acc_prev_year) if asset.dep_period_acc_prev_year else None,
            float(asset.dep_period_yearly) if asset.dep_period_yearly else None,
            float(asset.dep_period_until_year) if asset.dep_period_until_year else None,
            float(asset.dep_period_remain) if asset.dep_period_remain else None,
            float(asset.acc_depreciation_prev) if asset.acc_depreciation_prev else None,
            float(asset.net_book_value_prev) if asset.net_book_value_prev else None,
            float(asset.dep_expense_current) if asset.dep_expense_current else None,
            float(asset.acc_depreciation_curr) if asset.acc_depreciation_curr else None,
            float(asset.net_book_value_curr) if asset.net_book_value_curr else None,
            "Yes" if asset.status_additional else "No",
            "Yes" if asset.status_disposals else "No",
            asset.condition,
            asset.photo_status,
            asset.remark,
        ]
        apply_row(ws, ri, values, row_fill, num_cols)

    auto_col_width(ws)


def build_monthly_depreciation(wb: Workbook, assets: list, year_ref: int):
    ws = wb.create_sheet("Monthly Depreciation")
    ws.sheet_properties.tabColor = "1E3A8A"
    ws.freeze_panes = "B2"

    headers = ["Job / Category"] + MONTHS + ["Total/Year"]
    apply_header(ws, 1, headers, "1E3A8A")

    # Aggregate by job
    job_data: dict = {}
    for asset in assets:
        job = asset.job or "UNKNOWN"
        if job not in job_data:
            job_data[job] = {k: 0.0 for k in MONTH_KEYS}
            job_data[job]["total"] = 0.0
        for dm in asset.depreciation_monthly:
            if dm.year == year_ref:
                mk = MONTH_KEYS[dm.month - 1]
                amt = float(dm.amount or 0)
                job_data[job][mk] += amt
                job_data[job]["total"] += amt

    num_cols = set(range(2, 15))  # months + total

    grand = {k: 0.0 for k in MONTH_KEYS}
    grand["total"] = 0.0
    for ri, (job, data) in enumerate(sorted(job_data.items()), start=2):
        row_fill = "EFF6FF" if ri % 2 == 0 else None
        row_vals = [job] + [data[k] for k in MONTH_KEYS] + [data["total"]]
        apply_row(ws, ri, row_vals, row_fill, num_cols)
        for k in MONTH_KEYS:
            grand[k] += data[k]
        grand["total"] += data["total"]

    # Grand total row
    total_row = len(job_data) + 2
    grand_vals = ["GRAND TOTAL"] + [grand[k] for k in MONTH_KEYS] + [grand["total"]]
    apply_header(ws, total_row, grand_vals, "1E3A8A")

    auto_col_width(ws)


def build_by_group(wb: Workbook, assets: list, year_ref: int):
    ws = wb.create_sheet("Summary by Group")
    ws.sheet_properties.tabColor = "374151"
    ws.freeze_panes = "A2"

    headers = ["Asset Group", "Count", "Purchase Price",
               "Acc. Depreciation", "Net Book Value", f"Yearly Dep. {year_ref}"]
    apply_header(ws, 1, headers, "374151")
    num_cols = {3, 4, 5, 6}

    group_data: dict = {}
    for asset in assets:
        g = asset.group_name or "Unknown"
        if g not in group_data:
            group_data[g] = {"count": 0, "purchase_price": 0.0,
                             "acc_dep": 0.0, "nbv": 0.0, "yearly_dep": 0.0}
        group_data[g]["count"] += 1
        group_data[g]["purchase_price"] += float(asset.purchase_price or 0)
        group_data[g]["acc_dep"] += float(asset.acc_depreciation_curr or 0)
        group_data[g]["nbv"] += float(asset.net_book_value_curr or 0)
        # yearly dep from monthly rows
        for dm in asset.depreciation_monthly:
            if dm.year == year_ref:
                group_data[g]["yearly_dep"] += float(dm.amount or 0)

    totals = {"count": 0, "purchase_price": 0.0, "acc_dep": 0.0, "nbv": 0.0, "yearly_dep": 0.0}
    for ri, (grp, d) in enumerate(sorted(group_data.items()), start=2):
        row_fill = "F9FAFB" if ri % 2 == 0 else None
        values = [grp, d["count"], d["purchase_price"], d["acc_dep"], d["nbv"], d["yearly_dep"]]
        apply_row(ws, ri, values, row_fill, num_cols)
        for k in totals:
            totals[k] += d[k]

    total_row = len(group_data) + 2
    total_vals = ["TOTAL", totals["count"], totals["purchase_price"],
                  totals["acc_dep"], totals["nbv"], totals["yearly_dep"]]
    apply_header(ws, total_row, total_vals, "374151")
    for ci in num_cols:
        cell = ws.cell(row=total_row, column=ci)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal="right")

    auto_col_width(ws)


def build_by_category(wb: Workbook, assets: list, year_ref: int):
    ws = wb.create_sheet("Summary by Category")
    ws.sheet_properties.tabColor = "0F4C81"
    ws.freeze_panes = "A2"

    headers = ["Category", "Count", "Purchase Price",
               "Acc. Depreciation", "Net Book Value", f"Yearly Dep. {year_ref}", "% of Total"]
    apply_header(ws, 1, headers, "0F4C81")
    num_cols = {3, 4, 5, 6}

    cat_data: dict = {}
    for asset in assets:
        c = asset.category or "UNKNOWN"
        if c not in cat_data:
            cat_data[c] = {"count": 0, "purchase_price": 0.0,
                           "acc_dep": 0.0, "nbv": 0.0, "yearly_dep": 0.0}
        cat_data[c]["count"] += 1
        cat_data[c]["purchase_price"] += float(asset.purchase_price or 0)
        cat_data[c]["acc_dep"] += float(asset.acc_depreciation_curr or 0)
        cat_data[c]["nbv"] += float(asset.net_book_value_curr or 0)
        for dm in asset.depreciation_monthly:
            if dm.year == year_ref:
                cat_data[c]["yearly_dep"] += float(dm.amount or 0)

    grand_yearly = sum(d["yearly_dep"] for d in cat_data.values())
    totals = {"count": 0, "purchase_price": 0.0, "acc_dep": 0.0, "nbv": 0.0, "yearly_dep": 0.0}

    sorted_cats = sorted(cat_data.items(), key=lambda x: x[1]["yearly_dep"], reverse=True)
    for ri, (cat, d) in enumerate(sorted_cats, start=2):
        row_fill = "EFF6FF" if ri % 2 == 0 else None
        pct = (d["yearly_dep"] / grand_yearly * 100) if grand_yearly > 0 else 0
        values = [cat, d["count"], d["purchase_price"], d["acc_dep"], d["nbv"], d["yearly_dep"], round(pct, 2)]
        apply_row(ws, ri, values, row_fill, num_cols)
        # Percent cell
        pct_cell = ws.cell(row=ri, column=7)
        pct_cell.number_format = '0.00"%"'
        pct_cell.alignment = Alignment(horizontal="right")
        for k in totals:
            totals[k] += d[k]

    total_row = len(cat_data) + 2
    total_vals = ["TOTAL", totals["count"], totals["purchase_price"],
                  totals["acc_dep"], totals["nbv"], totals["yearly_dep"], "100%"]
    apply_header(ws, total_row, total_vals, "0F4C81")
    for ci in num_cols:
        cell = ws.cell(row=total_row, column=ci)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal="right")

    auto_col_width(ws)


def build_acquisitions(wb: Workbook, acquisitions: list, year_ref: int):
    ws = wb.create_sheet("Acquisitions & Disposals")
    ws.sheet_properties.tabColor = "047857"
    ws.freeze_panes = "A2"

    headers = ["ID", "Asset ID", "Site", "Job", "Fixed Asset No",
               "Application", "Transaction Date", "Bookslip No",
               "Asset Name", "Price", "Status", "Vendor/Customer", "Year Ref", "Created At"]
    apply_header(ws, 1, headers, "047857")
    num_cols = {10}

    for ri, acq in enumerate(acquisitions, start=2):
        row_fill = "F0FDF4" if ri % 2 == 0 else None
        values = [
            acq.id,
            acq.asset_id,
            acq.site,
            acq.job,
            acq.fixed_asset_no,
            acq.application,
            acq.transaction_date,
            acq.bookslip_no,
            acq.asset_name,
            float(acq.price) if acq.price else None,
            acq.status,
            acq.vendor_customer,
            acq.year_ref,
            acq.created_at,
        ]
        apply_row(ws, ri, values, row_fill, num_cols)

    auto_col_width(ws)


# ── Main endpoint ──────────────────────────────────────────────────────────────

@router.get("/excel")
def export_excel(
    year_ref: int = Query(2026, ge=2025),
    site_location: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Generate and download a multi-sheet Excel file with all depreciation data."""

    # Fetch assets with monthly data
    q = (
        db.query(FixedAsset)
        .options(joinedload(FixedAsset.depreciation_monthly))
        .filter(FixedAsset.year_ref == year_ref)
    )
    if site_location:
        q = q.filter(FixedAsset.site_location == site_location)
    assets = q.order_by(FixedAsset.no).all()

    # Fetch acquisitions
    aq = db.query(AcquisitionDisposal).filter(AcquisitionDisposal.year_ref == year_ref)
    if site_location:
        aq = aq.filter(AcquisitionDisposal.site == site_location)
    acquisitions = aq.order_by(AcquisitionDisposal.id).all()

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    build_asset_list(wb, assets, year_ref)
    build_monthly_depreciation(wb, assets, year_ref)
    build_by_group(wb, assets, year_ref)
    build_by_category(wb, assets, year_ref)
    build_acquisitions(wb, acquisitions, year_ref)

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    site_label = site_location or "All"
    filename = f"depreciation_{year_ref}_{site_label}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Forecast export helpers ────────────────────────────────────────────────────

def _months_elapsed_fc(purchase_date: date, target_year: int, target_month: int) -> int:
    sy, sm = purchase_date.year, purchase_date.month
    return max(0, (target_year - sy) * 12 + (target_month - sm) + 1)


def _calc_planned_fc(asset: PlannedAsset, forecast_year: int):
    if not asset.purchase_price or not asset.depreciation_period_total:
        return None
    price = float(asset.purchase_price)
    period = int(asset.depreciation_period_total)
    if period <= 0:
        return None
    monthly = price / period
    pd = date(asset.planned_purchase_year, asset.planned_purchase_month, 1)
    prev_year = forecast_year - 1
    acc_prev = min(_months_elapsed_fc(pd, prev_year, 12), period)
    acc_curr = min(_months_elapsed_fc(pd, forecast_year, 12), period)
    yearly = max(0, acc_curr - acc_prev)
    month_amounts = []
    for m in range(1, 13):
        elapsed = _months_elapsed_fc(pd, forecast_year, m)
        month_amounts.append(monthly if 0 < elapsed <= period else 0.0)
    total_year = sum(month_amounts)
    return {
        "name": asset.name,
        "group_name": asset.group_name or "",
        "category": asset.category or "",
        "site_location": asset.site_location or "",
        "job": asset.job or "",
        "purchase_price": price,
        "period": period,
        "purchase_month": asset.planned_purchase_month,
        "purchase_year": asset.planned_purchase_year,
        "monthly_dep": monthly,
        "yearly_dep": total_year,
        "nbv_curr": max(0.0, price - monthly * acc_curr),
        "months": month_amounts,
        "job_key": asset.job or "UNKNOWN",
        "group_key": asset.group_name or "Unknown",
        "category_key": asset.category or "UNKNOWN",
    }


def _calc_existing_fc(asset: FixedAsset, forecast_year: int):
    if not asset.purchase_price or not asset.depreciation_period_total:
        return None
    price = float(asset.purchase_price)
    period = int(asset.depreciation_period_total)
    if period <= 0:
        return None
    monthly = price / period
    prev_year = forecast_year - 1
    if asset.purchase_date:
        acc_prev = min(_months_elapsed_fc(asset.purchase_date, prev_year, 12), period)
        acc_curr = min(_months_elapsed_fc(asset.purchase_date, forecast_year, 12), period)
    else:
        acc_prev = acc_curr = 0
    month_amounts = []
    for m in range(1, 13):
        if asset.purchase_date:
            elapsed = _months_elapsed_fc(asset.purchase_date, forecast_year, m)
            month_amounts.append(monthly if 0 < elapsed <= period else 0.0)
        else:
            month_amounts.append(0.0)
    return {
        "job_key": asset.job or "UNKNOWN",
        "group_key": asset.group_name or "Unknown",
        "category_key": asset.category or "UNKNOWN",
        "months": month_amounts,
        "yearly_dep": sum(month_amounts),
        "purchase_price": price,
        "acc_dep": monthly * acc_curr,
        "nbv_curr": max(0.0, price - monthly * acc_curr),
    }


def build_forecast_planned(wb: Workbook, planned: list, forecast_year: int):
    ws = wb.create_sheet("Planned Assets")
    ws.sheet_properties.tabColor = "C05621"
    ws.freeze_panes = "A2"
    headers = [
        "Name", "Group", "Category", "Site", "Job",
        "Purchase Price", "Period (mo)", "Purchase Month", "Purchase Year",
        "Monthly Dep.",
        *[f"{m} {forecast_year}" for m in MONTHS],
        f"Total Dep. {forecast_year}", "NBV Projected",
    ]
    apply_header(ws, 1, headers, "C05621")
    num_cols = {6, 10, *range(11, 25), 24, 25}

    for ri, asset in enumerate(planned, start=2):
        fc = _calc_planned_fc(asset, forecast_year)
        if not fc:
            continue
        row_fill = "FFF7ED" if ri % 2 == 0 else None
        month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                       "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        values = [
            fc["name"], fc["group_name"], fc["category"],
            fc["site_location"], fc["job"],
            fc["purchase_price"], fc["period"],
            month_names[fc["purchase_month"] - 1], fc["purchase_year"],
            fc["monthly_dep"],
            *fc["months"],
            fc["yearly_dep"], fc["nbv_curr"],
        ]
        apply_row(ws, ri, values, row_fill, num_cols)

    # Footer totals
    if planned:
        total_row = len([p for p in planned if _calc_planned_fc(p, forecast_year)]) + 2
        fcs = [_calc_planned_fc(p, forecast_year) for p in planned if _calc_planned_fc(p, forecast_year)]
        total_vals = [
            f"TOTAL ({len(fcs)} planned assets)", "", "", "", "",
            sum(f["purchase_price"] for f in fcs), "", "", "",
            sum(f["monthly_dep"] for f in fcs),
            *[sum(f["months"][i] for f in fcs) for i in range(12)],
            sum(f["yearly_dep"] for f in fcs),
            sum(f["nbv_curr"] for f in fcs),
        ]
        apply_header(ws, total_row, total_vals, "C05621")
        for ci in num_cols:
            ws.cell(row=total_row, column=ci).number_format = '#,##0'
            ws.cell(row=total_row, column=ci).alignment = Alignment(horizontal="right")

    auto_col_width(ws)


def build_forecast_monthly(wb: Workbook, existing: list, planned: list, forecast_year: int):
    ws = wb.create_sheet("Forecast Monthly")
    ws.sheet_properties.tabColor = "1E3A8A"
    ws.freeze_panes = "B2"

    # Build per-job aggregation
    job_existing: dict = {}
    for asset in existing:
        fc = _calc_existing_fc(asset, forecast_year)
        if not fc:
            continue
        key = fc["job_key"]
        if key not in job_existing:
            job_existing[key] = [0.0] * 12
        for i in range(12):
            job_existing[key][i] += fc["months"][i]

    job_planned: dict = {}
    for asset in planned:
        fc = _calc_planned_fc(asset, forecast_year)
        if not fc:
            continue
        key = fc["job_key"]
        if key not in job_planned:
            job_planned[key] = [0.0] * 12
        for i in range(12):
            job_planned[key][i] += fc["months"][i]

    all_jobs = sorted(set(list(job_existing.keys()) + list(job_planned.keys())))

    headers = ["Job / Category"] + MONTHS + ["Total/Year"]
    apply_header(ws, 1, headers, "1E3A8A")
    num_cols = set(range(2, 15))

    ri = 2
    for job in all_jobs:
        ex = job_existing.get(job, [0.0] * 12)
        pl = job_planned.get(job, [0.0] * 12)
        combined = [ex[i] + pl[i] for i in range(12)]
        row_fill = "EFF6FF" if ri % 2 == 0 else None
        values = [job] + combined + [sum(combined)]
        apply_row(ws, ri, values, row_fill, num_cols)
        ri += 1
        # Sub-rows if there are planned amounts
        if any(pl):
            apply_row(ws, ri, [f"  ↳ {job} (planned)"] + pl + [sum(pl)], "FFF7ED", num_cols)
            ri += 1

    # Grand total
    grand_ex = [sum(job_existing.get(j, [0.0]*12)[i] for j in all_jobs) for i in range(12)]
    grand_pl = [sum(job_planned.get(j, [0.0]*12)[i] for j in all_jobs) for i in range(12)]
    grand_total = [grand_ex[i] + grand_pl[i] for i in range(12)]
    apply_header(ws, ri, ["GRAND TOTAL"] + grand_total + [sum(grand_total)], "1E3A8A")

    auto_col_width(ws)


def build_forecast_by_group(wb: Workbook, existing: list, planned: list, forecast_year: int):
    ws = wb.create_sheet("Forecast by Group")
    ws.sheet_properties.tabColor = "374151"
    ws.freeze_panes = "A2"

    headers = [
        "Asset Group",
        "Existing Assets", "Planned Assets",
        f"Existing Yearly Dep. {forecast_year}",
        f"Planned Yearly Dep. {forecast_year}",
        f"Total Yearly Dep. {forecast_year}",
        "Purchase Price (Existing)", "Acc. Dep.", "NBV Projected",
    ]
    apply_header(ws, 1, headers, "374151")
    num_cols = {4, 5, 6, 7, 8, 9}

    group_ex: dict = {}
    for asset in existing:
        fc = _calc_existing_fc(asset, forecast_year)
        if not fc:
            continue
        k = fc["group_key"]
        if k not in group_ex:
            group_ex[k] = {"count": 0, "yearly": 0.0, "price": 0.0, "acc": 0.0, "nbv": 0.0}
        group_ex[k]["count"] += 1
        group_ex[k]["yearly"] += fc["yearly_dep"]
        group_ex[k]["price"] += fc["purchase_price"]
        group_ex[k]["acc"] += fc["acc_dep"]
        group_ex[k]["nbv"] += fc["nbv_curr"]

    group_pl: dict = {}
    for asset in planned:
        fc = _calc_planned_fc(asset, forecast_year)
        if not fc:
            continue
        k = fc["group_key"]
        if k not in group_pl:
            group_pl[k] = {"count": 0, "yearly": 0.0}
        group_pl[k]["count"] += 1
        group_pl[k]["yearly"] += fc["yearly_dep"]

    all_groups = sorted(set(list(group_ex.keys()) + list(group_pl.keys())))
    totals = [0, 0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]

    for ri, grp in enumerate(all_groups, start=2):
        ex = group_ex.get(grp, {"count": 0, "yearly": 0.0, "price": 0.0, "acc": 0.0, "nbv": 0.0})
        pl = group_pl.get(grp, {"count": 0, "yearly": 0.0})
        row_fill = "F9FAFB" if ri % 2 == 0 else None
        total_dep = ex["yearly"] + pl["yearly"]
        values = [grp, ex["count"], pl["count"], ex["yearly"], pl["yearly"], total_dep,
                  ex["price"], ex["acc"], ex["nbv"]]
        apply_row(ws, ri, values, row_fill, num_cols)
        for i, v in enumerate([ex["count"], pl["count"], ex["yearly"], pl["yearly"], total_dep,
                                ex["price"], ex["acc"], ex["nbv"]]):
            totals[i] += v

    total_row = len(all_groups) + 2
    total_vals = ["TOTAL"] + totals
    apply_header(ws, total_row, total_vals, "374151")
    for ci in num_cols:
        ws.cell(row=total_row, column=ci).number_format = '#,##0'
        ws.cell(row=total_row, column=ci).alignment = Alignment(horizontal="right")

    auto_col_width(ws)


def build_forecast_by_category(wb: Workbook, existing: list, planned: list, forecast_year: int):
    ws = wb.create_sheet("Forecast by Category")
    ws.sheet_properties.tabColor = "0F4C81"
    ws.freeze_panes = "A2"

    headers = [
        "Category",
        "Existing Assets", "Planned Assets",
        f"Existing Yearly Dep. {forecast_year}",
        f"Planned Yearly Dep. {forecast_year}",
        f"Total Yearly Dep. {forecast_year}",
        "% of Total",
    ]
    apply_header(ws, 1, headers, "0F4C81")
    num_cols = {4, 5, 6}

    cat_ex: dict = {}
    for asset in existing:
        fc = _calc_existing_fc(asset, forecast_year)
        if not fc:
            continue
        k = fc["category_key"]
        if k not in cat_ex:
            cat_ex[k] = {"count": 0, "yearly": 0.0}
        cat_ex[k]["count"] += 1
        cat_ex[k]["yearly"] += fc["yearly_dep"]

    cat_pl: dict = {}
    for asset in planned:
        fc = _calc_planned_fc(asset, forecast_year)
        if not fc:
            continue
        k = fc["category_key"]
        if k not in cat_pl:
            cat_pl[k] = {"count": 0, "yearly": 0.0}
        cat_pl[k]["count"] += 1
        cat_pl[k]["yearly"] += fc["yearly_dep"]

    all_cats = sorted(set(list(cat_ex.keys()) + list(cat_pl.keys())))
    grand_total = sum((cat_ex.get(c, {}).get("yearly", 0.0) + cat_pl.get(c, {}).get("yearly", 0.0))
                      for c in all_cats)

    for ri, cat in enumerate(sorted(all_cats, key=lambda c: -(
        cat_ex.get(c, {}).get("yearly", 0.0) + cat_pl.get(c, {}).get("yearly", 0.0)
    )), start=2):
        ex = cat_ex.get(cat, {"count": 0, "yearly": 0.0})
        pl = cat_pl.get(cat, {"count": 0, "yearly": 0.0})
        total_dep = ex["yearly"] + pl["yearly"]
        pct = round(total_dep / grand_total * 100, 2) if grand_total > 0 else 0.0
        row_fill = "EFF6FF" if ri % 2 == 0 else None
        values = [cat, ex["count"], pl["count"], ex["yearly"], pl["yearly"], total_dep, pct]
        apply_row(ws, ri, values, row_fill, num_cols)
        ws.cell(row=ri, column=7).number_format = '0.00"%"'
        ws.cell(row=ri, column=7).alignment = Alignment(horizontal="right")

    total_row = len(all_cats) + 2
    total_vals = [
        "TOTAL",
        sum(cat_ex.get(c, {}).get("count", 0) for c in all_cats),
        sum(cat_pl.get(c, {}).get("count", 0) for c in all_cats),
        sum(cat_ex.get(c, {}).get("yearly", 0.0) for c in all_cats),
        sum(cat_pl.get(c, {}).get("yearly", 0.0) for c in all_cats),
        grand_total, "100%",
    ]
    apply_header(ws, total_row, total_vals, "0F4C81")
    for ci in num_cols:
        ws.cell(row=total_row, column=ci).number_format = '#,##0'
        ws.cell(row=total_row, column=ci).alignment = Alignment(horizontal="right")

    auto_col_width(ws)


@router.get("/forecast-excel")
def export_forecast_excel(
    forecast_year: int = Query(2027, ge=2026),
    site_location: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Generate and download a multi-sheet forecast Excel file."""
    from app.routers.forecast import get_planned_assets

    # Fetch existing assets (base year 2026)
    q = db.query(FixedAsset).filter(FixedAsset.year_ref == 2026)
    if site_location:
        q = q.filter(FixedAsset.site_location == site_location)
    existing = q.all()

    # Fetch planned assets (covers forecast_year)
    planned = get_planned_assets(db, forecast_year, site_location)

    wb = Workbook()
    wb.remove(wb.active)

    build_forecast_planned(wb, planned, forecast_year)
    build_forecast_monthly(wb, existing, planned, forecast_year)
    build_forecast_by_group(wb, existing, planned, forecast_year)
    build_forecast_by_category(wb, existing, planned, forecast_year)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    site_label = site_location or "All"
    filename = f"forecast_{forecast_year}_{site_label}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
