"""
Excel Importer: imports data from Est-Depreciation_Calculation_for_2026_.xlsx
into the PostgreSQL database.

Usage:
    python -m app.utils.excel_importer
    python -m app.utils.excel_importer --file /path/to/file.xlsx
"""
import argparse
import sys
import os
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Optional

import openpyxl
from sqlalchemy.orm import Session

# Add parent dir to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

from app.database import SessionLocal, engine
from app.models import FixedAsset, DepreciationMonthly, AcquisitionDisposal
from app.database import Base

MONTHS = [None, "jan", "feb", "mar", "apr", "may", "jun",
          "jul", "aug", "sep", "oct", "nov", "dec"]


def to_decimal(value) -> Optional[Decimal]:
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def to_int(value) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(float(str(value)))
    except (ValueError, TypeError):
        return None


def to_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def to_str(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def import_fa_sheet(ws, year_ref: int, db: Session) -> int:
    """Import assets from FA_2025 or FA_2026 sheet. Returns count of imported rows."""
    count = 0
    for row in ws.iter_rows(min_row=26, values_only=True):
        # Col 0 = No (must be numeric)
        if row[0] is None or not isinstance(row[0], (int, float)):
            continue
        no = to_int(row[0])

        asset_no = to_str(row[5])
        if not asset_no:
            continue

        name = to_str(row[10])
        if not name:
            name = asset_no  # fallback

        # Monthly depreciation values for Jan-Dec (cols 29-40)
        monthly_values = {}
        for month_idx, col_idx in enumerate(range(29, 41), start=1):
            if col_idx < len(row):
                monthly_values[month_idx] = to_decimal(row[col_idx])

        data = dict(
            no=no,
            site_location=to_str(row[1]),
            job=to_str(row[2]),
            account_no=to_int(row[3]),
            category=to_str(row[4]),
            asset_no=asset_no,
            fixed_asset_number_ax=to_str(row[6]),
            purchase_date=to_date(row[7]),
            group_name=to_str(row[8]),
            voucher_no=to_str(row[9]),
            name=name,
            maker_type_location=to_str(row[11]),
            capacity_size_user=to_str(row[12]),
            year=to_int(row[13]),
            police_no=to_str(row[14]),
            machine_no=to_str(row[15]),
            chasis_no=to_str(row[16]),
            quantity=to_int(row[17]) or 1,
            valas=to_str(row[18]),
            purchase_price=to_decimal(row[19]),
            monthly_depreciation=to_decimal(row[20]),
            depreciation_period_total=to_int(row[21]),
            dep_period_acc_prev_year=to_decimal(row[22]),
            dep_period_yearly=to_decimal(row[23]),
            dep_period_until_year=to_decimal(row[24]),
            dep_period_remain=to_decimal(row[25]),
            acc_depreciation_prev=to_decimal(row[26]),
            net_book_value_prev=to_decimal(row[27]),
            dep_expense_current=to_decimal(row[28]),
            acc_depreciation_curr=to_decimal(row[41]) if len(row) > 41 else None,
            net_book_value_curr=to_decimal(row[42]) if len(row) > 42 else None,
            status_additional=bool(row[43]) if len(row) > 43 and row[43] else False,
            status_disposals=bool(row[44]) if len(row) > 44 and row[44] else False,
            condition=to_str(row[45]) if len(row) > 45 else None,
            photo_status=to_str(row[47]) if len(row) > 47 else None,
            remark=to_str(row[46]) if len(row) > 46 else None,
            year_ref=year_ref,
        )

        # Upsert: update if exists, insert if not
        existing = db.query(FixedAsset).filter(FixedAsset.asset_no == asset_no).first()
        if existing:
            for k, v in data.items():
                setattr(existing, k, v)
            asset = existing
        else:
            asset = FixedAsset(**data)
            db.add(asset)

        db.flush()

        # Insert monthly depreciation rows
        db.query(DepreciationMonthly).filter(
            DepreciationMonthly.asset_id == asset.id,
            DepreciationMonthly.year == year_ref,
        ).delete()

        for month, amount in monthly_values.items():
            dm = DepreciationMonthly(
                asset_id=asset.id,
                year=year_ref,
                month=month,
                amount=amount or Decimal("0"),
            )
            db.add(dm)

        count += 1

    return count


def import_approval_sheet(ws, year_ref: int, db: Session) -> int:
    """Import acquisition/disposal records from List Approval sheets."""
    count = 0
    for row in ws.iter_rows(min_row=6, values_only=True):
        if row[0] is None or not isinstance(row[0], (int, float)):
            continue

        asset_name = to_str(row[9])
        if not asset_name:
            continue

        fixed_asset_no = to_str(row[5]) if len(row) > 5 else None

        # Try to find linked asset
        asset_id = None
        if fixed_asset_no:
            asset = db.query(FixedAsset).filter(
                FixedAsset.asset_no == fixed_asset_no
            ).first()
            if asset:
                asset_id = asset.id

        record = AcquisitionDisposal(
            asset_id=asset_id,
            site=to_str(row[1]),
            job=to_str(row[2]),
            fixed_asset_no=fixed_asset_no,
            application=to_str(row[6]) if len(row) > 6 else None,
            transaction_date=to_date(row[7]) if len(row) > 7 else None,
            bookslip_no=to_str(row[8]) if len(row) > 8 else None,
            asset_name=asset_name,
            price=to_decimal(row[10]) if len(row) > 10 else None,
            status=to_str(row[11]) if len(row) > 11 else None,
            vendor_customer=to_str(row[13]) if len(row) > 13 else None,
            year_ref=year_ref,
        )
        db.add(record)
        count += 1

    return count


def run_import(xlsx_path: str):
    print(f"Opening: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    print(f"Sheets found: {wb.sheetnames}")

    # Create tables if they don't exist
    Base.metadata.create_all(bind=engine)

    db: Session = SessionLocal()
    try:
        total = 0

        # Import FA_2026
        if "FA_2026" in wb.sheetnames:
            n = import_fa_sheet(wb["FA_2026"], 2026, db)
            print(f"  FA_2026: {n} assets imported")
            total += n

        # Import FA_2025
        if "FA_2025" in wb.sheetnames:
            n = import_fa_sheet(wb["FA_2025"], 2025, db)
            print(f"  FA_2025: {n} assets imported")
            total += n

        db.commit()

        # Import List Approval sheets
        for sheet_name in ["List Approval 2022", "List Approval 2021"]:
            if sheet_name in wb.sheetnames:
                year = int(sheet_name.split()[-1])
                n = import_approval_sheet(wb[sheet_name], year, db)
                print(f"  {sheet_name}: {n} records imported")
                db.commit()

        print(f"\nDone. Total assets: {total}")

    except Exception as e:
        db.rollback()
        print(f"ERROR: {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import Excel data into database")
    parser.add_argument(
        "--file",
        default=r"E:\Asset-depre\Est-Depreciation_Calculation_for_2026_.xlsx",
        help="Path to the Excel file",
    )
    args = parser.parse_args()
    run_import(args.file)
