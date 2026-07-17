"""
Excel Importer V2 - Auto-detect format, supports multiple file versions.

Supports:
- File lama: Est-Depreciation_Calculation_for_2026_.xlsx (header row 23-24, data row 26+)
- File baru: Fixed Asset_202606.xlsx (header row 7-8, data row 10+, no Category column)
- Future files: auto-detects header row by scanning for keyword markers

Category derivation (when Category column is absent):
  Derived from site_location + job using known mapping patterns.
"""
import argparse
import io
import os
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Optional

import openpyxl
from sqlalchemy.orm import Session

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

from app.database import SessionLocal, engine
from app.models import AcquisitionDisposal, DepreciationMonthly, FixedAsset
from app.database import Base

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def to_decimal(value) -> Optional[Decimal]:
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def to_int(value) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(float(str(value)))
    except (ValueError, TypeError):
        return None


def to_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def to_str(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# Category derivation
# ---------------------------------------------------------------------------

def derive_category(site_location: Optional[str], job: Optional[str]) -> Optional[str]:
    """
    Derive the Category field from site_location and/or job when the
    Category column is not present in the source file.

    Mapping is based on patterns observed in the file lama data.
    """
    site = (site_location or '').upper().strip()
    job_up = (job or '').upper().strip()

    # Site-based rules (most specific)
    if 'BALI' in site:
        return 'BALI'
    if site in ('MDN', 'MEDAN') or 'MDN' in site:
        return 'MDN'
    if site in ('PBG', 'PALEMBANG') or 'PBG' in site:
        return 'PBG'
    if site in ('MKS', 'MAKASSAR') or 'MKS' in site:
        return 'MKS'
    if site in ('SBY', 'SURABAYA') or 'SBY' in site:
        return 'SBY'

    # Job-based rules
    if 'CLCTRP' in job_up or 'TRP' in job_up:
        return 'TRP'
    if 'CLCWH' == job_up:
        return 'WH1'
    if 'WH2' in job_up:
        return 'WH2'
    if 'WH' in job_up:
        return 'WH1'
    if 'CLCOWH' in job_up:
        return 'OTHERS'

    # Fallback: use site_location directly if short code
    if site and len(site) <= 5:
        return site

    return 'OTHERS'


# ---------------------------------------------------------------------------
# Auto-detect file format
# ---------------------------------------------------------------------------

# Keywords that must ALL appear in the header row for it to be identified
HEADER_KEYWORDS = {'no.', 'name of fixed asset', 'purchase price', 'fixed asset number'}


def detect_header_row(ws) -> int:
    """
    Scan rows 1-40 and return the first row number that contains all
    HEADER_KEYWORDS (case-insensitive). Returns -1 if not found.
    """
    for i in range(1, 41):
        row = list(ws.iter_rows(min_row=i, max_row=i, values_only=True))[0]
        texts = {str(v).strip().lower() for v in row if v is not None}
        # Also check combined text (some headers span two rows)
        combined = ' '.join(texts)
        if all(kw in combined for kw in HEADER_KEYWORDS):
            return i
    return -1


def build_col_map(header_row: tuple, sub_row: tuple) -> dict:
    """
    Build a mapping of field-name → column-index by scanning the header
    and sub-header rows. Returns a dict with known field names as keys.
    """
    col_map = {}
    max_col = max(len(header_row), len(sub_row))

    def norm(v):
        return str(v).strip().lower().replace('\n', ' ').replace('  ', ' ') if v else ''

    for i in range(max_col):
        h1 = norm(header_row[i] if i < len(header_row) else None)
        h2 = norm(sub_row[i] if i < len(sub_row) else None)
        combined = f"{h1} {h2}".strip()

        if h1 == 'no.' and 'no.' not in col_map:
            col_map['no'] = i
        elif h1 in ('site/', 'site/\nlocation', 'site/ location', 'site/location'):
            col_map['site_location'] = i
        elif h1 == 'job':
            col_map['job'] = i
        elif h1 == 'account':
            col_map['account_no'] = i
        elif h1 == 'category':
            col_map['category'] = i
        elif h1 == 'asset' and h2 == 'no.' and 'asset_no' not in col_map:
            col_map['asset_no'] = i
        elif h1 == 'fixed asset number' and h2 == 'ax':
            col_map['fixed_asset_number_ax'] = i
        elif h1 == 'purchase date':
            col_map['purchase_date'] = i
        elif h1 == 'group':
            col_map['group_name'] = i
        elif h1 == 'voucher':
            col_map['voucher_no'] = i
        elif h1 == 'name of fixed asset':
            col_map['name'] = i
        elif h1 == 'maker/type/location':
            col_map['maker_type_location'] = i
        elif h1 == 'capacity/size/user':
            col_map['capacity_size_user'] = i
        elif h1 == 'year':
            col_map['year'] = i
        elif h1 == 'police' and h2 == 'no.':
            col_map['police_no'] = i
        elif h1 == 'machine' and h2 == 'no.':
            col_map['machine_no'] = i
        elif h1 == 'chasis' and h2 == 'no.':
            col_map['chasis_no'] = i
        elif h1 == 'quantity':
            col_map['quantity'] = i
        elif h1 == 'valas':
            col_map['valas'] = i
        elif h1 == 'purchase price':
            col_map['purchase_price'] = i
        elif h1 == 'monthly depreciation':
            col_map['monthly_depreciation'] = i
        elif 'depreciation period' in h1 and h2 == 'total':
            col_map['dep_period_total'] = i
        elif 'acc.' in h2 and ('2025' in h2 or '2024' in h2 or '2023' in h2):
            col_map['dep_period_acc_prev'] = i
        elif 'yearly' in h2:
            col_map['dep_period_yearly'] = i
        elif 'until' in h2:
            col_map['dep_period_until'] = i
        elif 'remain' in h2:
            col_map['dep_period_remain'] = i
        elif h1 == 'accumulated depreciation' and 'acc_depr_prev' not in col_map:
            col_map['acc_depr_prev'] = i
        elif h1 == 'net book value' and 'nbv_prev' not in col_map:
            col_map['nbv_prev'] = i
        elif 'depreciation expense' in h1:
            col_map['dep_expense'] = i
        elif h1 == 'depreciation' and h2 != '' and 'monthly' in combined:
            # DEPRECIATION 2026 (MONTHLY) — Jan starts here
            col_map['monthly_jan'] = i
        elif h2 == 'jan' and 'monthly_jan' not in col_map:
            col_map['monthly_jan'] = i
        elif h2 == 'feb' and 'monthly_jan' in col_map:
            pass  # consecutive from jan
        elif h1 == 'accumulated depreciation' and 'acc_depr_curr' not in col_map and 'acc_depr_prev' in col_map:
            col_map['acc_depr_curr'] = i
        elif h1 == 'net book value' and 'nbv_curr' not in col_map and 'nbv_prev' in col_map:
            col_map['nbv_curr'] = i
        elif h1 == 'status' and h2 == 'additional':
            col_map['status_additional'] = i
        elif h2 == 'disposals' and 'status_additional' in col_map:
            col_map['status_disposals'] = i
        elif h1 == 'acquisition' and h2 == 'condition':
            col_map['condition'] = i
        elif h1 == 'remark':
            col_map['remark'] = i
        elif 'photo' in h1 and 'status' in h1:
            col_map['photo_status'] = i

    return col_map


def resolve_monthly_cols(sub_row: tuple, jan_col: int) -> dict:
    """
    Given the sub-header row and the column index of 'Jan', map month numbers
    1-12 to column indices by scanning forward for month name matches.
    """
    month_names = ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
                   'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
    month_map = {}
    # Scan from jan_col forward for up to 20 columns
    mi = 0
    for ci in range(jan_col, min(jan_col + 20, len(sub_row))):
        cell = str(sub_row[ci]).strip().lower() if sub_row[ci] else ''
        if mi < 12 and cell == month_names[mi]:
            month_map[mi + 1] = ci
            mi += 1
    return month_map


# ---------------------------------------------------------------------------
# Main importer
# ---------------------------------------------------------------------------

def import_fa_sheet(ws, year_ref: int, db: Session) -> int:
    """
    Auto-detect format and import assets from a FA sheet.
    Returns count of upserted rows.
    """
    # 1. Find header row
    header_row_idx = detect_header_row(ws)
    if header_row_idx == -1:
        print(f"  WARNING: Could not detect header row in sheet. Skipping.")
        return 0

    sub_row_idx = header_row_idx + 1
    data_start_row = header_row_idx + 3  # skip header, sub-header, formula row

    rows = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))
    sub_rows = list(ws.iter_rows(min_row=sub_row_idx, max_row=sub_row_idx, values_only=True))
    header_tuple = rows[0] if rows else ()
    sub_tuple = sub_rows[0] if sub_rows else ()

    col_map = build_col_map(header_tuple, sub_tuple)
    has_category = 'category' in col_map

    # Find monthly columns
    jan_col = col_map.get('monthly_jan')
    if jan_col is None:
        # fallback: scan sub_row for 'jan'
        for ci, v in enumerate(sub_tuple):
            if v and str(v).strip().lower() == 'jan':
                jan_col = ci
                col_map['monthly_jan'] = ci
                break

    monthly_cols = resolve_monthly_cols(sub_tuple, jan_col) if jan_col is not None else {}

    # Resolve acc_depr_curr and nbv_curr from sub_row dates
    # They appear after the monthly columns
    if jan_col and monthly_cols:
        last_month_col = max(monthly_cols.values()) if monthly_cols else jan_col + 11
        for ci in range(last_month_col + 1, min(last_month_col + 5, len(sub_tuple))):
            v = sub_tuple[ci] if ci < len(sub_tuple) else None
            if v is not None and 'acc_depr_curr' not in col_map:
                col_map['acc_depr_curr'] = ci
            elif v is not None and 'nbv_curr' not in col_map and 'acc_depr_curr' in col_map:
                col_map['nbv_curr'] = ci
                break

    # Determine acc_depr_prev and nbv_prev from the columns before dep_expense
    # They appear right after dep_period_remain and before dep_expense
    if 'dep_period_remain' in col_map and 'acc_depr_prev' not in col_map:
        remain_col = col_map['dep_period_remain']
        col_map['acc_depr_prev'] = remain_col + 1
        col_map['nbv_prev'] = remain_col + 2
        col_map['dep_expense'] = remain_col + 3

    count = 0

    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        # Row must start with a numeric No.
        no_col = col_map.get('no', 0)
        if no_col >= len(row):
            continue
        if row[no_col] is None or not isinstance(row[no_col], (int, float)):
            continue

        def get(field, default=None):
            ci = col_map.get(field)
            if ci is None or ci >= len(row):
                return default
            return row[ci]

        asset_no = to_str(get('asset_no'))
        if not asset_no or asset_no == '-':
            asset_no = None

        ax_no = to_str(get('fixed_asset_number_ax'))
        if not ax_no or ax_no == '-':
            ax_no = None

        # Must have at least AX number to import
        if not ax_no:
            continue  # skip aset tanpa AX number

        # Use AX number as asset_no fallback if not provided
        if not asset_no:
            asset_no = ax_no

        name = to_str(get('name')) or asset_no

        # Category: from column only. If file has no Category column, leave None.
        # Never derive from site_location — user inputs category manually.
        if has_category:
            category = to_str(get('category'))
        else:
            category = None  # preserve existing DB value on upsert

        data = dict(
            no=to_int(get('no')),
            site_location=to_str(get('site_location')),
            job=to_str(get('job')),
            account_no=to_int(get('account_no')),
            category=category,
            asset_no=asset_no,
            fixed_asset_number_ax=ax_no,  # already validated above
            purchase_date=to_date(get('purchase_date')),
            group_name=to_str(get('group_name')),
            voucher_no=to_str(get('voucher_no')),
            name=name,
            maker_type_location=to_str(get('maker_type_location')),
            capacity_size_user=to_str(get('capacity_size_user')),
            year=to_int(get('year')),
            police_no=to_str(get('police_no')),
            machine_no=to_str(get('machine_no')),
            chasis_no=to_str(get('chasis_no')),
            quantity=to_int(get('quantity')) or 1,
            valas=to_str(get('valas')),
            purchase_price=to_decimal(get('purchase_price')),
            monthly_depreciation=to_decimal(get('monthly_depreciation')),
            depreciation_period_total=to_int(get('dep_period_total')),
            dep_period_acc_prev_year=to_decimal(get('dep_period_acc_prev')),
            dep_period_yearly=to_decimal(get('dep_period_yearly')),
            dep_period_until_year=to_decimal(get('dep_period_until')),
            dep_period_remain=to_decimal(get('dep_period_remain')),
            acc_depreciation_prev=to_decimal(get('acc_depr_prev')),
            net_book_value_prev=to_decimal(get('nbv_prev')),
            dep_expense_current=to_decimal(get('dep_expense')),
            acc_depreciation_curr=to_decimal(get('acc_depr_curr')),
            net_book_value_curr=to_decimal(get('nbv_curr')),
            status_additional=bool(get('status_additional')) if get('status_additional') else False,
            status_disposals=bool(get('status_disposals')) if get('status_disposals') else False,
            condition=to_str(get('condition')),
            photo_status=to_str(get('photo_status')),
            remark=to_str(get('remark')),
            year_ref=year_ref,
        )

        # Upsert by fixed_asset_number_ax + year_ref (ax_no already validated, never None)
        existing = db.query(FixedAsset).filter(
            FixedAsset.fixed_asset_number_ax == ax_no,
            FixedAsset.year_ref == year_ref,
        ).first()

        if existing:
            for k, v in data.items():
                # Preserve existing category if new file has no Category column
                if k == 'category' and not has_category and existing.category:
                    continue
                setattr(existing, k, v)
            asset = existing
        else:
            asset = FixedAsset(**data)
            db.add(asset)

        db.flush()

        # Monthly depreciation rows
        db.query(DepreciationMonthly).filter(
            DepreciationMonthly.asset_id == asset.id,
            DepreciationMonthly.year == year_ref,
        ).delete()

        for month, ci in monthly_cols.items():
            if ci < len(row):
                amount = to_decimal(row[ci]) or Decimal("0")
                db.add(DepreciationMonthly(
                    asset_id=asset.id,
                    year=year_ref,
                    month=month,
                    amount=amount,
                ))

        count += 1

    return count


def import_approval_sheet(ws, year_ref: int, db: Session) -> int:
    """Import acquisition/disposal records from 'List Approval' sheet.

    Actual column layout (0-indexed) — confirmed from Fixed Asset_202606_1.xlsx:
      0:  No.
      1:  Site
      2:  Job
      3:  Fixed Asset Number AX
      4:  Asset No. (internal, e.g. CLC-A-002)
      5:  Application No (e.g. 001/ASC/FA-Disp/I/2025)
      6:  (empty / sub-number)
      7:  (empty / sub-number)
      8:  Remark / Date (secondary date — used as transaction_date if col9 absent)
      9:  Purchase / Disposed Date (primary date)
      10: Bookslip No.
      11: Name of Fixed Asset
      12: Purchase / Disposal Price
      13: Status — Acquisition column (value: 'Acquisition' or None)
      14: Status — Disposals column (value: 'Disposal'/'Disposals' or None)
      15: Vendor / Customer
    """
    count = 0

    # Find header row — look for 'site' and 'job' keywords
    data_start = 7  # safe default
    for r in range(1, 15):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        texts = [str(v).strip().lower() for v in row if v]
        if 'site' in texts and 'job' in texts:
            data_start = r + 2  # skip header row + one blank row
            break

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        # Skip rows with no data
        if not row or len(row) < 4:
            continue
        # Row must have a non-empty site or job to be a valid data row
        site = to_str(row[1]) if len(row) > 1 else None
        job = to_str(row[2]) if len(row) > 2 else None
        if not site and not job:
            continue

        fixed_asset_no_ax = to_str(row[3]) if len(row) > 3 else None
        application = to_str(row[5]) if len(row) > 5 else None

        # col 9 = Purchase/Disposed Date (primary), col 8 = Remark date (fallback)
        transaction_date = to_date(row[9]) if len(row) > 9 else None
        if transaction_date is None:
            transaction_date = to_date(row[8]) if len(row) > 8 else None

        bookslip_no = to_str(row[10]) if len(row) > 10 else None
        asset_name = to_str(row[11]) if len(row) > 11 else None
        price = to_decimal(row[12]) if len(row) > 12 else None

        # Status: col 13 = 'Acquisition', col 14 = 'Disposal'/'Disposals'
        status_acq = to_str(row[13]) if len(row) > 13 else None
        status_dis = to_str(row[14]) if len(row) > 14 else None
        status = status_acq or status_dis

        # col 15 = Vendor/Customer
        vendor_customer = to_str(row[15]) if len(row) > 15 else None

        if not asset_name and not fixed_asset_no_ax:
            continue

        # Try to find linked asset by AX number first, then asset_no
        asset_id = None
        if fixed_asset_no_ax:
            asset = db.query(FixedAsset).filter(
                FixedAsset.fixed_asset_number_ax == fixed_asset_no_ax
            ).first()
            if not asset:
                asset = db.query(FixedAsset).filter(
                    FixedAsset.asset_no == fixed_asset_no_ax
                ).first()
            if asset:
                asset_id = asset.id

        record = AcquisitionDisposal(
            asset_id=asset_id,
            site=site,
            job=job,
            fixed_asset_no=fixed_asset_no_ax,
            application=application,
            transaction_date=transaction_date,
            bookslip_no=bookslip_no,
            asset_name=asset_name,
            price=price,
            status=status,
            vendor_customer=vendor_customer,
            year_ref=year_ref,
        )
        db.add(record)
        count += 1

    return count


def run_import(xlsx_path: str, wb=None):
    """
    Main entry point. Accepts either a file path (str) or a pre-loaded
    openpyxl workbook (for use from FastAPI endpoint).
    """
    if wb is None:
        print(f"Opening: {xlsx_path}")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    print(f"Sheets: {wb.sheetnames}")

    # Detect year_ref from sheet date or filename
    year_ref = 2026  # default

    # Ensure tables exist
    Base.metadata.create_all(bind=engine)

    db: Session = SessionLocal()
    results = {}
    try:
        # Import FA sheets
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith('FA_'):
                try:
                    yr = int(sheet_name.split('_')[1])
                except (IndexError, ValueError):
                    yr = year_ref
                n = import_fa_sheet(wb[sheet_name], yr, db)
                results[sheet_name] = n
                print(f"  {sheet_name}: {n} assets imported (year_ref={yr})")
                db.commit()

        # Import List Approval / Disposal sheets
        for sheet_name in wb.sheetnames:
            sn_lower = sheet_name.lower()
            if 'list approval' in sn_lower or 'disposal' in sn_lower:
                # Extract year from sheet name
                import re
                years_found = re.findall(r'20\d{2}', sheet_name)
                yr = int(years_found[-1]) if years_found else year_ref
                n = import_approval_sheet(wb[sheet_name], yr, db)
                results[sheet_name] = n
                print(f"  {sheet_name}: {n} records imported")
                db.commit()

        total_assets = sum(v for k, v in results.items() if k.startswith('FA_'))
        print(f"\nDone. Total assets: {total_assets}")
        return results

    except Exception as e:
        db.rollback()
        print(f"ERROR: {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import Excel (auto-detect format)")
    parser.add_argument(
        "--file",
        default=r"E:\Asset-depre\Est-Depreciation_Calculation_for_2026_.xlsx",
        help="Path to the Excel file",
    )
    args = parser.parse_args()
    run_import(args.file)
